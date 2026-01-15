from __future__ import annotations

import csv
import os
import random
import re
import time
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None


def _normalize_newlines(value: str) -> str:
    return value.replace("\r\n", "\n").replace("\r", "\n")


def _split_filename_keywords(value: str) -> List[str]:
    if not value:
        return []
    parts = [kw.strip() for kw in value.split("、") if kw.strip()]
    return parts


def _match_filename_keywords(filename: str, keywords: List[str], match_mode: str) -> bool:
    if not keywords:
        return True
    normalized = filename
    if match_mode == "精准匹配":
        stem = Path(filename).stem
        return all(kw == stem for kw in keywords)
    return all(kw in normalized for kw in keywords)


def _is_separator_line(line: str) -> bool:
    stripped = line.strip()
    if not stripped:
        return False
    parts = stripped.split()
    if len(parts) < 10:
        return False
    return all(part in {"◆", "◇"} for part in parts)


def _ensure_parent_dir(path: Path) -> None:
    if not path.parent.exists():
        path.parent.mkdir(parents=True, exist_ok=True)


def _sanitize_filename(value: str) -> str:
    name = (value or "").strip()
    if not name:
        return ""

    # Windows/NTFS illegal characters + control chars
    name = re.sub(r"[<>:\"/\\|?*\x00-\x1F]", "_", name)
    name = name.strip(" .")
    if not name:
        return ""

    # Avoid Windows reserved device names
    base = name.split(".", 1)[0].upper()
    reserved = {
        "CON",
        "PRN",
        "AUX",
        "NUL",
        *(f"COM{i}" for i in range(1, 10)),
        *(f"LPT{i}" for i in range(1, 10)),
    }
    if base in reserved:
        name = f"_{name}"

    return name[:200]


def _split_terms(value: str) -> List[str]:
    if not value:
        return []
    parts = re.split(r"[、\n]+", value)
    return [part.strip() for part in parts if part.strip()]


def _split_custom_terms(
    value: str,
    delimiter: str,
    trim: bool,
    remove_empty: bool,
    allow_newline: bool = True,
) -> List[str]:
    if not value:
        return []
    delimiter = delimiter or "、"
    normalized = _normalize_newlines(value)
    pattern = rf"(?:{re.escape(delimiter)}|\n)+" if allow_newline else re.escape(delimiter)
    raw_parts = re.split(pattern, normalized)
    results: List[str] = []
    for part in raw_parts:
        entry = part.strip() if trim else part
        if remove_empty and not entry.strip():
            continue
        results.append(entry)
    if remove_empty:
        return [item for item in results if item.strip()]
    return results


def _string_to_list(value: Any) -> List[str]:
    if isinstance(value, list):
        return [str(item) for item in value]
    if isinstance(value, str):
        return _normalize_newlines(value).split("\n")
    return [str(value)]


def _deduplicate_keep_order(items: Iterable[str]) -> List[str]:
    seen = set()
    result: List[str] = []
    for item in items:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return result


def _split_story_lines(story: str) -> List[str]:
    return [line.strip() for line in _normalize_newlines(story or "").split("\n") if line.strip()]


def _extract_dialogues(story: str) -> List[Tuple[str, str]]:
    dialogues: List[Tuple[str, str]] = []
    for raw_line in _split_story_lines(story):
        if "：" in raw_line:
            speaker, content = raw_line.split("：", 1)
            speaker = speaker.strip() or "旁白"
            content = content.strip()
            if content:
                dialogues.append((speaker, content))
        else:
            dialogues.append(("旁白", raw_line))
    return dialogues


def _collect_character_names(dialogues: List[Tuple[str, str]]) -> List[str]:
    names: List[str] = []
    for speaker, _ in dialogues:
        if speaker not in names:
            names.append(speaker)
    return names


def _infer_character_traits(dialogues: List[Tuple[str, str]]) -> Dict[str, Dict[str, str]]:
    traits: Dict[str, Dict[str, str]] = {}
    for speaker, line in dialogues:
        entry = traits.setdefault(speaker, {"出场": 0, "关键词": set()})
        entry["出场"] += 1
        for keyword in re.findall(r"[\u4e00-\u9fa5]{2,}", line):
            if len(keyword) <= 4:
                entry["关键词"].add(keyword)
    summary: Dict[str, Dict[str, str]] = {}
    for speaker, info in traits.items():
        keywords = list(info["关键词"])[:5]
        summary[speaker] = {
            "出场次数": str(info["出场"]),
            "印象关键词": "、".join(keywords) if keywords else "待补充",
        }
    return summary


def _infer_scenes(story: str) -> List[str]:
    location_candidates = [
        "街道", "办公室", "咖啡馆", "学校", "医院", "公园", "夜市", "酒吧",
        "沙漠", "森林", "海边", "山顶", "地铁站", "客厅", "车站", "古城",
        "雨夜", "清晨", "黄昏", "舞台", "仓库", "市场", "后巷",
    ]
    scenes: List[str] = []
    text = _normalize_newlines(story or "")
    for keyword in location_candidates:
        if keyword in text:
            scenes.append(keyword)
    if not scenes:
        scenes.extend(["开场场景", "冲突场景", "高潮场景", "结尾场景"])
    return scenes[:6]


def _build_shot_outline(story: str, character_info: str, scene_info: str) -> List[Dict[str, str]]:
    base_scenes = _split_story_lines(scene_info)
    scenes = base_scenes if base_scenes else _infer_scenes(story)
    characters = _split_story_lines(character_info)
    shots: List[Dict[str, str]] = []
    for idx, scene in enumerate(scenes, start=1):
        focus = characters[(idx - 1) % len(characters)] if characters else "主要角色"
        shots.append(
            {
                "镜头": f"第{idx}镜头",
                "场景": scene,
                "重点": focus,
                "摄影": "中景" if idx % 2 == 0 else "近景",
                "提示": "突出情绪张力" if idx >= len(scenes) - 1 else "交代情节走向",
            }
        )
    return shots


def _format_shot_script(shots: List[Dict[str, str]]) -> str:
    lines: List[str] = []
    for shot in shots:
        lines.append(
            (
                f"{shot['镜头']}｜场景：{shot['场景']}｜主角焦点：{shot['重点']}｜摄影：{shot['摄影']}"
                f"｜导演提示：{shot['提示']}"
            )
        )
    return "\n".join(lines)


_NUMBER_PREFIX_PATTERNS = [
    re.compile(r"^\s*\d{1,4}[.\-_、）)]\s*"),
    re.compile(r"^\s*\d{1,4}\s+"),
    re.compile(r"^\s*[（(]?\d+[）)]\s*"),
    re.compile(r"^\s*[一二三四五六七八九十百千]+[、.)）]\s*"),
]


def _strip_number_prefix(line: str) -> str:
    for pattern in _NUMBER_PREFIX_PATTERNS:
        if pattern.match(line):
            return pattern.sub("", line, count=1).lstrip()
    return line


def _parse_keywords_field(value: str, delimiter: str = "") -> List[str]:
    """把关键字字段解析为关键字列表。支持逗号、顿号、分号或换行作为分隔符，或使用自定义分隔符。"""
    if not value:
        return []
    text = str(value)
    if delimiter and delimiter.strip():
        parts = re.split(re.escape(delimiter), text)
    else:
        parts = re.split(r"[，,;\n]+", text)
    return [p.strip() for p in parts if p.strip()]


def _try_extract_prefix(line: str) -> Tuple[Optional[str], str]:
    """尝试提取行首的编号/前缀，返回 (prefix, rest)。若无前缀，prefix 返回 None，rest 返回原行。"""
    if not line:
        return None, line
    # 常见前缀样式示例: 1., 1a., 1a)、(1)、1-、(1a)、一. 等
    m = re.match(r"^\s*([0-9]+[a-zA-Z]{0,2}|[一二三四五六七八九十]+)[.\-、\)）\]]+\s*(.*)$", line)
    if m:
        return m.group(1), m.group(2)
    # 备用：像 1a. 或 1.a
    m2 = re.match(r"^\s*([0-9]+[a-zA-Z]{1,2})[.\-]\s*(.*)$", line)
    if m2:
        return m2.group(1), m2.group(2)
    return None, line


def _compile_regex_list(patterns: List[str], case_sensitive: bool) -> List[re.Pattern]:
    compiled: List[re.Pattern] = []
    flags = 0 if case_sensitive else re.IGNORECASE
    for p in patterns:
        try:
            compiled.append(re.compile(p, flags))
        except re.error:
            # 忽略无法编译的正则，调用处应捕获并报告
            continue
    return compiled


def _line_matches_keywords(
    line: str,
    keywords: List[str],
    match_mode: str,
    match_logic: str,
    case_sensitive: bool,
) -> Tuple[bool, List[str]]:
    """判断一行是否匹配关键字集合。返回 (matched, matched_keywords_list)。"""
    if not keywords:
        return (False, [])

    text = line if case_sensitive else line.lower()
    normalized_keywords = [k if case_sensitive else k.lower() for k in keywords]

    matched_keywords: List[str] = []

    if match_mode == "正则匹配":
        regexes = _compile_regex_list(keywords, case_sensitive)
        for idx, rg in enumerate(regexes):
            if rg.search(line):
                matched_keywords.append(keywords[idx])
        matched = False
        if match_logic == "AND":
            matched = len(matched_keywords) == len(regexes) and len(regexes) > 0
        else:
            matched = len(matched_keywords) > 0
        return matched, matched_keywords

    for kw in normalized_keywords:
        if match_mode == "包含匹配":
            if kw in text:
                matched_keywords.append(kw)
        elif match_mode == "前缀匹配":
            # 匹配行首（忽略前导空格）
            if text.lstrip().startswith(kw):
                matched_keywords.append(kw)
        elif match_mode == "精确匹配":
            if text.strip() == kw:
                matched_keywords.append(kw)
        else:
            # 退化为包含匹配
            if kw in text:
                matched_keywords.append(kw)

    if match_logic == "AND":
        return (len(matched_keywords) == len(normalized_keywords) and len(normalized_keywords) > 0, matched_keywords)
    return (len(matched_keywords) > 0, matched_keywords)


def _parse_index_list(text: str) -> List[int]:
    indices: List[int] = []
    if not text or not text.strip():
        return indices
    parts = re.split(r"[，,;\n]", text)
    for raw in parts:
        part = raw.strip()
        if not part:
            continue
        if "-" in part:
            start_str, end_str = part.split("-", 1)
            start = int(start_str)
            end = int(end_str)
            step = 1 if end >= start else -1
            indices.extend(range(start, end + step, step))
        else:
            value = int(part)
            indices.append(value)
    ordered_unique: List[int] = []
    seen = set()
    for idx in indices:
        if idx not in seen:
            seen.add(idx)
            ordered_unique.append(idx)
    return ordered_unique


def _ensure_table_dimensions(headers: List[str], rows: List[List[Any]], min_rows: int) -> None:
    column_count = len(headers)
    for row in rows:
        if len(row) < column_count:
            row.extend([""] * (column_count - len(row)))
        elif len(row) > column_count:
            del row[column_count:]
    while len(rows) < min_rows:
        rows.append(["" for _ in headers])


def _table_to_matrix(headers: List[str], data_rows: List[Dict[str, Any]]) -> List[List[Any]]:
    if not headers:
        return []
    return [[row.get(header, "") for header in headers] for row in data_rows]


def _matrix_to_dict_rows(headers: List[str], matrix: List[List[Any]]) -> List[Dict[str, Any]]:
    dict_rows: List[Dict[str, Any]] = []
    for row in matrix:
        dict_rows.append({header: ("" if value is None else value) for header, value in zip(headers, row)})
    return dict_rows


def _read_csv_table(path: Path, header_row: int) -> Tuple[List[str], List[Dict[str, Any]]]:
    with path.open("r", encoding="utf-8", newline="") as fp:
        reader = csv.reader(fp)
        rows = list(reader)
    if not rows:
        return [], []
    headers = rows[header_row]
    data_rows = rows[header_row + 1 :]
    table = [dict(zip(headers, row)) for row in data_rows if any(cell != "" for cell in row)]
    return headers, table


def _read_xlsx_table(path: Path, sheet_name: Optional[str], header_row: int) -> Tuple[List[str], List[Dict[str, Any]]]:
    if load_workbook is None:
        raise RuntimeError("需要 openpyxl 才能读取 XLSX 表格")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.values)
    if not rows:
        return [], []
    headers = [str(cell) if cell is not None else "" for cell in rows[header_row]]
    table: List[Dict[str, Any]] = []
    for raw_row in rows[header_row + 1 :]:
        if raw_row is None:
            continue
        values = ["" if cell is None else cell for cell in raw_row]
        if not any(val != "" for val in values):
            continue
        table.append(dict(zip(headers, values)))
def _load_workbook_with_sheet(path: Path, sheet_name: Optional[str]):
    if Workbook is None:
        raise RuntimeError("需要 openpyxl 才能写入 XLSX 表格")
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
    try:
        if sheet_name:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
        else:
            ws = wb.active
        yield wb, ws
    finally:
        wb.save(path)
        wb.close()

class _WorkbookWithContext:
    """工作簿上下文管理器"""
    def __init__(self, wb: Workbook):
        self.wb = wb
    
    def __enter__(self):
        return self.wb
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.close()


class buding_简易文本替换:
    CATEGORY = "Buding/TextStory/文本处理"
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("替换后文本",)
    FUNCTION = "replace"

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "source_text": ("STRING", {"default": "", "multiline": True, "tooltip": "输入待替换的原始文本"}),
                "original_terms": ("STRING", {"default": "", "multiline": True, "tooltip": "需查找的原文，可用顿号或换行分隔"}),
                "replacement_terms": ("STRING", {"default": "", "multiline": True, "tooltip": "对应替换词，可用顿号或换行分隔"}),
                "case_sensitive": ("BOOLEAN", {"default": True, "tooltip": "开启后区分大小写"}),
                "replace_all": ("BOOLEAN", {"default": True, "tooltip": "开启后替换所有匹配，关闭仅替换第一次"}),
            }
        }

    def replace(
        self,
        source_text: str,
        original_terms: str,
        replacement_terms: str,
        case_sensitive: bool,
        replace_all: bool,
    ):
        text = source_text or ""
        originals = _split_terms(original_terms)
        replacements = _split_terms(replacement_terms)
        pair_count = min(len(originals), len(replacements))
        for index in range(pair_count):
            pattern = originals[index]
            replacement = replacements[index]
            if not pattern:
                continue
            if case_sensitive:
                if replace_all:
                    text = text.replace(pattern, replacement)
                else:
                    text = text.replace(pattern, replacement, 1)
            else:
                escaped = re.escape(pattern)
                compiled = re.compile(escaped, re.IGNORECASE)
                if replace_all:
                    text = compiled.sub(replacement, text)
                else:
                    text = compiled.sub(replacement, text, count=1)
        return (text,)


class buding_符号文本转换:
    CATEGORY = "Buding/TextStory/文本处理"
    RETURN_TYPES = ("STRING", "INT")
    RETURN_NAMES = ("换行文本", "条目数量")
    FUNCTION = "convert"

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "source_text": ("STRING", {"default": "", "multiline": True, "tooltip": "输入包含分隔符的文本"}),
                "delimiter": ("STRING", {"default": "、", "tooltip": "拆分用的分隔符，可自定义多个字符"}),
                "trim_entries": ("BOOLEAN", {"default": True, "tooltip": "去除每项首尾空格"}),
                "remove_empty": ("BOOLEAN", {"default": True, "tooltip": "过滤空白条目"}),
            }
        }

    def convert(self, source_text: str, delimiter: str, trim_entries: bool, remove_empty: bool):
        items = _split_custom_terms(source_text, delimiter, trim_entries, remove_empty, allow_newline=True)
        newline_text = "\n".join(items)
        return (newline_text, len(items))


class buding_列文本转换:
    CATEGORY = "Buding/TextStory/文本处理"
    RETURN_TYPES = ("STRING", "INT")
    RETURN_NAMES = ("符号文本", "条目数量")
    FUNCTION = "convert"

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "source_text": ("STRING", {"default": "", "multiline": True, "tooltip": "输入按行排列的文本"}),
                "trim_entries": ("BOOLEAN", {"default": True, "tooltip": "去除每项首尾空格"}),
                "remove_empty": ("BOOLEAN", {"default": True, "tooltip": "过滤空白行"}),
            }
        }

    def convert(self, source_text: str, trim_entries: bool, remove_empty: bool):
        items = _split_custom_terms(source_text, "\n", trim_entries, remove_empty, allow_newline=False)
        delimiter_text = "、".join(items)
        return (delimiter_text, len(items))


class buding_自动文本转换:
    CATEGORY = "Buding/TextStory/文本处理"
    RETURN_TYPES = ("STRING", "INT", "STRING")
    RETURN_NAMES = ("转换结果", "条目数量", "使用模式")
    FUNCTION = "convert"

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "source_text": ("STRING", {"default": "", "multiline": True, "tooltip": "自动识别并转换的文本"}),
                "delimiter": ("STRING", {"default": "、", "tooltip": "当判定为符号模式时使用的分隔符"}),
                "trim_entries": ("BOOLEAN", {"default": True, "tooltip": "去除每项首尾空格"}),
                "remove_empty": ("BOOLEAN", {"default": True, "tooltip": "过滤空白条目"}),
            }
        }

    def convert(self, source_text: str, delimiter: str, trim_entries: bool, remove_empty: bool):
        text = source_text or ""
        if not text.strip():
            return ("", 0, "未检测到内容")

        normalized = _normalize_newlines(text)
        delimiter = delimiter or "、"
        has_delimiter = delimiter in normalized
        has_newline = "\n" in normalized

        delimiter_count = normalized.count(delimiter)
        newline_count = normalized.count("\n")

        if has_delimiter and (delimiter_count >= newline_count or not has_newline):
            items = _split_custom_terms(text, delimiter, trim_entries, remove_empty, allow_newline=True)
            converted = "\n".join(items)
            mode = "符号转行"
        else:
            items = _split_custom_terms(text, "\n", trim_entries, remove_empty, allow_newline=False)
            converted = delimiter.join(items)
            mode = "行转符号"

        return (converted, len(items), mode)


class buding_读取TXT文件:
    CATEGORY = "Buding/TextStory/文本处理"
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("全文本",)
    FUNCTION = "read"

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "file_path": ("STRING", {"default": "output/文本保存/示例", "tooltip": "TXT 文件路径或文件夹路径"}),
                "scan_depth": ("INT", {"default": 0, "min": 0, "max": 10, "tooltip": "扫描深度：0=仅当前目录，1=含1层子目录"}),
                "filename_keyword": ("STRING", {"default": "", "tooltip": "文件名关键词过滤"}),
                "filename_match_mode": (["包含匹配", "精准匹配"], {"default": "包含匹配", "tooltip": "关键词匹配模式：包含=只要关键词出现在文件名中即可；精准=仅在文件名（不含扩展）完全等于关键词时才匹配"}),
                "strip_empty_lines": ("BOOLEAN", {"default": True, "tooltip": "是否移除空行"}),
                "deduplicate": ("BOOLEAN", {"default": False, "tooltip": "是否去除重复行"}),
                "filter_header_mode": (["不开启", "模式1：第一个空行", "模式2：最后一个空行"], {"default": "不开启"}),
                "filter_separator": ("BOOLEAN", {"default": False, "tooltip": "开启后仅输出最后两条分隔行（◆ ◇ ◆）之间的内容"}),
                "always_reload": ("BOOLEAN", {"default": False, "tooltip": "开启后始终重新读取"}),
            }
        }

    @classmethod
    def IS_CHANGED(cls, always_reload=False, **kwargs):
        if always_reload:
            return float("nan")
        return hash(frozenset(kwargs.items()))

    def read(self, file_path: str, scan_depth: int, filename_keyword: str,
             filename_match_mode: str, strip_empty_lines: bool, deduplicate: bool,
             filter_header_mode: str, filter_separator: bool, always_reload: bool):
        path = Path(file_path).expanduser()
        
        files = []
        if path.is_file():
            files = [path]
        elif path.is_dir():
            files = self._scan_txt_files(path, filename_keyword, scan_depth, match_mode=filename_match_mode)
        else:
            # 尝试补全 .txt
            if path.with_suffix(".txt").is_file():
                files = [path.with_suffix(".txt")]
            else:
                return ("",)

        all_content = []
        for f in sorted(files):
            try:
                content = f.read_text(encoding="utf-8")
                normalized = content.replace("\r\n", "\n").replace("\r", "\n")
                
                lines = normalized.split("\n")
                
                # 分隔行过滤逻辑 (优先级最高，先处理这个)
                if filter_separator:
                    sep_idxs = [i for i, line in enumerate(lines) if _is_separator_line(line)]
                    if len(sep_idxs) >= 2:
                        start = sep_idxs[-2] + 1
                        end = sep_idxs[-1]
                        lines = lines[start:end]
                    else:
                        lines = []

                # 标题过滤逻辑
                empty_indices = [i for i, line in enumerate(lines) if line.strip() == ""]
                if filter_header_mode == "模式1：第一个空行" and empty_indices:
                    lines = lines[empty_indices[0] + 1:]
                elif filter_header_mode == "模式2：最后一个空行" and empty_indices:
                    lines = lines[empty_indices[-1] + 1:]
                
                if strip_empty_lines:
                    lines = [line.strip() for line in lines if line.strip()]
                
                all_content.extend(lines)
            except Exception:
                continue

        if deduplicate:
            seen = set()
            unique_lines = []
            for line in all_content:
                if line not in seen:
                    seen.add(line)
                    unique_lines.append(line)
            all_content = unique_lines

        output_text = "\n".join(all_content)
        return (output_text,)

    def _scan_txt_files(self, directory: Path, keyword: str = "", max_depth: int = 0,
                        current_depth: int = 0, match_mode: str = "包含匹配") -> List[Path]:
        files = []
        if not directory.exists() or not directory.is_dir():
            return files
        try:
            for item in directory.iterdir():
                if item.is_file() and item.suffix.lower() == ".txt":
                    if keyword:
                        keywords = _split_filename_keywords(keyword)
                        if _match_filename_keywords(item.name, keywords, match_mode):
                            files.append(item)
                    else:
                        files.append(item)
                elif item.is_dir() and current_depth < max_depth:
                    files.extend(self._scan_txt_files(item, keyword, max_depth, current_depth + 1))
        except Exception:
            pass
        return files


class buding_读取表格数据:
    CATEGORY = "Buding/TextStory/表格处理"
    RETURN_TYPES = ("STRING", "INT", "BWF_TEXT_LIST", "INT")
    RETURN_NAMES = ("输出文本", "段落数量", "段落列表", "使用种子")
    FUNCTION = "read_table"

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "file_path": ("STRING", {"default": "output/表格保存/示例", "tooltip": "CSV/XLSX 文件路径，可不写扩展名"}),
                "sheet_name": ("STRING", {"default": "", "tooltip": "工作表名称，留空使用第一个"}),
                "row_range": ("STRING", {"default": "", "tooltip": "行范围，例如 1-10,15，留空表示全部"}),
                "column_range": ("STRING", {"default": "", "tooltip": "列范围，例如 1-3,5，留空表示全部"}),
                "use_title_match": ("BOOLEAN", {"default": False, "tooltip": "是否按首行标题模糊匹配列"}),
                "title_keywords": ("STRING", {"default": "", "tooltip": "标题匹配关键字，逗号或顿号分隔"}),
                "single_selected_index": ("INT", {"default": 1, "min": 1, "tooltip": "单段模式下选择第几段"}),
                "enable_random": ("BOOLEAN", {"default": False, "tooltip": "开启后启用随机输出"}),
                "randomize": ("BOOLEAN", {"default": True, "tooltip": "随机模式下自动生成种子"}),
                "seed": ("INT", {"default": 0, "tooltip": "自定义随机种子"}),
                "enable_full_text": ("BOOLEAN", {"default": False, "tooltip": "开启后输出多段文本"}),
                "full_output_limit": ("INT", {"default": 10, "min": 1, "tooltip": "多段模式下最多输出段落数量"}),
            }
        }

    def _load_table(self, file_path: str, sheet_name: str) -> Tuple[List[str], List[List[Any]]]:
        path = Path(file_path).expanduser()
        if path.suffix.lower() not in {".csv", ".xlsx", ".xlsm"}:
            path = path.with_suffix(".xlsx")
        if not path.exists():
            raise FileNotFoundError(f"未找到文件: {path}")

        suffix = path.suffix.lower()
        if suffix == ".csv":
            headers, dict_rows = _read_csv_table(path, 0)
        else:
            headers, dict_rows = _read_xlsx_table(path, sheet_name or None, 0)

        matrix = _table_to_matrix(headers, dict_rows)
        return headers, matrix

    def _select_columns(
        self,
        headers: List[str],
        column_range: str,
        use_title_match: bool,
        title_keywords: str,
    ) -> List[int]:
        total = len(headers)
        base_indices = list(range(total))
        specified = _parse_index_list(column_range)
        if specified:
            base_indices = [idx - 1 for idx in specified if 0 < idx <= total]

        if use_title_match and title_keywords.strip():
            keywords = [kw.strip().lower() for kw in re.split(r"[，,]", title_keywords) if len(kw.strip()) >= 2]
            if keywords:
                matched: List[int] = []
                for idx in base_indices:
                    header_lower = str(headers[idx]).lower()
                    if any(kw in header_lower for kw in keywords):
                        matched.append(idx)
                base_indices = matched
        return [idx for idx in base_indices if 0 <= idx < total]

    def _select_rows(self, matrix: List[List[Any]], row_range: str) -> List[int]:
        total = len(matrix)
        specified = _parse_index_list(row_range)
        if not specified:
            return list(range(total))
        rows = [idx - 1 for idx in specified if 0 < idx <= total]
        return rows

    def _collect_paragraphs(self, matrix: List[List[Any]], column_indices: List[int], row_indices: List[int]) -> List[str]:
        paragraphs: List[str] = []
        for col_idx in column_indices:
            for row_idx in row_indices:
                if row_idx < 0 or row_idx >= len(matrix):
                    continue
                row = matrix[row_idx]
                if col_idx >= len(row):
                    continue
                value = row[col_idx]
                if value is None:
                    continue
                text = str(value).strip()
                if text:
                    paragraphs.append(text)
        return paragraphs

    def read_table(
        self,
        file_path: str,
        sheet_name: str,
        row_range: str,
        column_range: str,
        use_title_match: bool,
        title_keywords: str,
        single_selected_index: int,
        enable_random: bool,
        randomize: bool,
        seed: int,
        enable_full_text: bool,
        full_output_limit: int,
    ):
        headers, matrix = self._load_table(file_path, sheet_name)
        if not headers:
            return ("", 0, [], 0)

        column_indices = self._select_columns(headers, column_range, use_title_match, title_keywords)
        if not column_indices:
            column_indices = list(range(len(headers)))

        row_indices = self._select_rows(matrix, row_range)
        if not row_indices:
            row_indices = list(range(len(matrix)))

        paragraphs = self._collect_paragraphs(matrix, column_indices, row_indices)
        if not paragraphs:
            return ("", 0, [], 0)

        used_seed = 0
        rng: Optional[random.Random] = None
        if enable_random:
            if randomize or seed == 0:
                used_seed = random.randint(1, 2_147_483_647)
            else:
                used_seed = seed
            rng = random.Random(used_seed)

        if not enable_full_text:
            if rng is not None:
                chosen = rng.choice(paragraphs)
            else:
                index = max(1, single_selected_index)
                if index > len(paragraphs):
                    index = len(paragraphs)
                chosen = paragraphs[index - 1]
            return (chosen, 1, [chosen], used_seed)

        limit = min(max(1, full_output_limit), len(paragraphs))
        if rng is not None:
            shuffled = paragraphs[:]
            rng.shuffle(shuffled)
            selected_paragraphs = shuffled[:limit]
        else:
            selected_paragraphs = paragraphs[:limit]
        combined = "\n".join(selected_paragraphs)
        return (combined, len(selected_paragraphs), selected_paragraphs, used_seed)


class buding_文本保存TXT:
    CATEGORY = "Buding/TextStory/文本处理"
    RETURN_TYPES = ("STRING", "BOOLEAN", "STRING")
    RETURN_NAMES = ("文件路径", "写入成功", "保存内容预览")
    FUNCTION = "save"

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "text": ("STRING", {"default": "", "multiline": True, "tooltip": "需要保存的文本内容"}),
                "target_path": ("STRING", {"default": "output/文本保存/示例", "tooltip": "目标文件完整路径，可不写扩展名"}),
                "save_filename": (
                    "STRING",
                    {
                        "default": "",
                        "multiline": False,
                        "tooltip": "保存文件名（可选）。留空则使用 target_path 中的文件名规则。可不写扩展名。",
                    },
                ),
                "header_title": (
                    "STRING",
                    {
                        "default": "",
                        "multiline": False,
                        "tooltip": "首行标题内容（可选）。留空则不写首行标题。",
                    },
                ),
                "add_separator_line": (
                    "BOOLEAN",
                    {
                        "default": False,
                        "tooltip": "开启则在写入内容末尾追加一行分隔符（标题/正文之后）。",
                    },
                ),
                "mode": (
                    ["覆盖写入", "追加写入", "自动顺延"],
                    {
                        "default": "覆盖写入",
                        "tooltip": "覆盖写入: 覆盖已有文件\n追加写入: 在文件末尾追加\n自动顺延: 自动递增序号保存 (如: 1102001.txt, 1102002.txt)",
                    },
                ),
                "strip_number_prefix": ("BOOLEAN", {"default": False, "tooltip": "移除每行的编号前缀"}),
                "prepend_newline": ("BOOLEAN", {"default": False, "tooltip": "追加写入前是否补两个换行"}),
                "auto_create_dirs": ("BOOLEAN", {"default": True, "tooltip": "目录不存在时自动创建"}),
            }
        }

    def save(
        self,
        text: str,
        target_path: str,
        save_filename: str,
        header_title: str,
        add_separator_line: bool,
        mode: str,
        strip_number_prefix: bool,
        prepend_newline: bool,
        auto_create_dirs: bool,
    ):
        path = Path(target_path).expanduser()

        # 若用户提供文件名：使用 target_path 作为“目录或文件路径”，最终写入该目录下的该文件名
        sanitized_name = _sanitize_filename(save_filename)
        if sanitized_name:
            if target_path.strip().endswith(("/", "\\")):
                base_dir = path
            elif path.exists() and path.is_dir():
                base_dir = path
            else:
                base_dir = path.parent

            chosen = Path(sanitized_name)
            if chosen.suffix == "":
                chosen = chosen.with_suffix(".txt")
            path = (base_dir / chosen.name).expanduser()
        
        # 处理自动顺延模式
        if mode == "自动顺延":
            if path.suffix == "":
                path = path.with_suffix(".txt")
            
            if auto_create_dirs:
                _ensure_parent_dir(path)
            
            # 获取自动顺延的文件路径
            path = self._get_auto_increment_path(path)
        else:
            # 原有的逻辑
            if path.suffix == "":
                path = path.with_suffix(".txt")
            if auto_create_dirs:
                _ensure_parent_dir(path)
        
        raw_text = text or ""
        normalized = _normalize_newlines(raw_text)
        lines = normalized.split("\n")
        if strip_number_prefix:
            lines = [_strip_number_prefix(line) for line in lines]
        processed = "\n".join(lines)

        separator_line = (
            "◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆ ◇ ◆"
        )

        parts: List[str] = []
        header_line = _normalize_newlines(header_title or "").replace("\n", " ").strip()
        if header_line:
            parts.append(header_line)
        if processed:
            parts.append(processed)
        if add_separator_line:
            parts.append(separator_line)
        processed = "\n".join(parts)
        
        # 兼容旧版本
        legacy_mode_map = {
            "overwrite": "覆盖写入",
            "append": "追加写入",
        }
        normalized_mode = legacy_mode_map.get(mode, mode or "覆盖写入")

        # 自动顺延模式总是覆盖写入（因为文件名已经是唯一的）
        if normalized_mode == "自动顺延":
            file_mode = "w"
        elif normalized_mode == "追加写入":
            if prepend_newline and processed and path.exists() and path.stat().st_size > 0:
                # 补到“两个换行”：根据文件尾部已有换行数量补齐
                trailing_newlines = 0
                try:
                    with path.open("rb") as fp:
                        fp.seek(-2, os.SEEK_END)
                        tail = fp.read(2)
                    for b in reversed(tail):
                        if b in (10, 13):  # \n or \r
                            trailing_newlines += 1
                        else:
                            break
                except OSError:
                    trailing_newlines = 0

                missing = max(0, 2 - trailing_newlines)
                if missing:
                    processed = ("\n" * missing) + processed
            file_mode = "a"
        else:  # 覆盖写入
            file_mode = "w"
        
        try:
            if normalized_mode not in {"覆盖写入", "追加写入", "自动顺延"}:
                raise ValueError("mode 仅支持 覆盖写入、追加写入 或 自动顺延")
            
            with path.open(file_mode, encoding="utf-8") as fp:
                fp.write(processed)
            
            # 生成保存内容预览
            preview = processed if processed else "（空内容）"
            # 如果内容太长，截取前500个字符并添加提示
            if len(preview) > 500:
                preview = preview[:500] + "\n\n...（内容过长，已截取前500字符）"
            
            return (str(path), True, preview)
        except Exception as exc:  # pragma: no cover
            return (f"写入失败: {exc}", False, f"写入失败: {exc}")
    
    def _get_auto_increment_path(self, base_path: Path) -> Path:
        """
        获取自动顺延的文件路径
        例如: output/故事/1102.txt -> output/故事/1102001.txt, 1102002.txt, ...
        """
        import glob
        
        directory = base_path.parent
        base_name = base_path.stem  # 不包含扩展名的文件名，如 "1102"
        
        # 查找已存在的文件: 1102001.txt, 1102002.txt, ...
        pattern = directory / f"{base_name}*.txt"
        existing_files = glob.glob(str(pattern))
        
        # 提取序号
        max_index = 0
        for file in existing_files:
            filename = Path(file).name
            # 从 "1102001.txt" 提取 "001"
            if filename.startswith(base_name) and filename.endswith('.txt'):
                number_part = filename[len(base_name):-4]  # 去掉前缀和.txt
                if number_part.isdigit():
                    max_index = max(max_index, int(number_part))
        
        # 下一个序号，格式化为3位数字
        next_index = max_index + 1
        new_filename = f"{base_name}{next_index:03d}.txt"
        
        return directory / new_filename


class _WorkbookSheetManager:
    """工作表管理器上下文管理器"""
    def __init__(self, path: Path, sheet_mode: str, sheet_name: str):
        self.path = path
        self.sheet_mode = sheet_mode
        self.sheet_name = sheet_name
        self.wb = None
        self.ws = None
    
    def __enter__(self):
        if Workbook is None:
            raise RuntimeError("需要 openpyxl 才能写入 XLSX 表格")
        
        if self.path.exists():
            self.wb = load_workbook(self.path)
        else:
            self.wb = Workbook()
            # 删除默认工作表，因为我们要创建新的
            if "Sheet" in self.wb.sheetnames:
                self.wb.remove(self.wb["Sheet"])
        
        if self.sheet_mode == "新建内页":
            # 新建内页模式
            if self.sheet_name and self.sheet_name.strip():
                # 使用指定名称
                new_sheet_name = self.sheet_name.strip()
                # 确保名称唯一
                counter = 1
                original_name = new_sheet_name
                while new_sheet_name in self.wb.sheetnames:
                    new_sheet_name = f"{original_name}_{counter}"
                    counter += 1
                self.ws = self.wb.create_sheet(title=new_sheet_name)
            else:
                # 自动生成名称: Sheet1, Sheet2, ...
                sheet_counter = len(self.wb.sheetnames) + 1
                new_sheet_name = f"Sheet{sheet_counter}"
                # 确保名称唯一
                while new_sheet_name in self.wb.sheetnames:
                    sheet_counter += 1
                    new_sheet_name = f"Sheet{sheet_counter}"
                self.ws = self.wb.create_sheet(title=new_sheet_name)
            print(f"✅ 创建新工作表: {self.ws.title}")
        else:
            # 使用现有模式
            if self.sheet_name and self.sheet_name.strip():
                # 使用指定工作表
                self.ws = self.wb[self.sheet_name] if self.sheet_name in self.wb.sheetnames else self.wb.create_sheet(title=self.sheet_name)
            else:
                # 使用第一个工作表
                self.ws = self.wb.active if self.wb.active else self.wb.create_sheet(title="Sheet1")
        
        return self.wb, self.ws
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.wb:
            self.wb.save(self.path)
            self.wb.close()

class buding_写入表格数据:
    CATEGORY = "Buding/TextStory/表格处理"
    RETURN_TYPES = ("STRING", "INT", "STRING")
    RETURN_NAMES = ("写入文件路径", "写入行数", "写入文本")
    FUNCTION = "write_column"

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "file_path": ("STRING", {"default": "output/表格保存/示例", "tooltip": "目标 CSV/XLSX 路径，可不写扩展名"}),
                "sheet_mode": (["使用现有", "新建内页"], {
                    "default": "使用现有",
                    "tooltip": "使用现有: 在指定工作表中写入\n新建内页: 创建全新工作表并写入数据"
                }),
                "sheet_name": ("STRING", {"default": "", "tooltip": "XLSX 工作表名称\n• 使用现有模式: 指定要写入的工作表，留空使用第一个\n• 新建内页模式: 新工作表的名称，留空自动生成"}),
                "text_content": ("STRING", {"default": "", "multiline": True, "tooltip": "按行写入的文本内容"}),
                "target_column": ("STRING", {"default": "1", "tooltip": "目标列编号（从 1 开始）或标题关键字"}),
                "write_mode": (
                    ["覆盖重写", "空位追加"],
                    {
                        "default": "覆盖重写",
                        "tooltip": "覆盖重写: 覆盖指定列的所有数据\n空位追加: 从空位开始追加数据",
                    },
                ),
                "enable_new_column": ("BOOLEAN", {"default": False, "tooltip": "开启后在目标列右侧新增一列"}),
                "new_column_title": ("STRING", {"default": "新列", "tooltip": "新增列模式下的列标题"}),
            }
        }

    def _resolve_column_index(self, headers: List[str], column_spec: str) -> Tuple[int, str]:
        tokens = [token.strip() for token in re.split(r"[，,]", column_spec or "") if token.strip()]
        title_candidates = [token for token in tokens if not token.isdigit()]
        numeric_candidates = [int(token) for token in tokens if token.isdigit()]

        if title_candidates and headers:
            lowered_headers = [str(h).lower() for h in headers]
            for keyword in title_candidates:
                keyword_lower = keyword.lower()
                for idx, header_lower in enumerate(lowered_headers):
                    if keyword_lower in header_lower:
                        return idx, headers[idx]

        for value in numeric_candidates:
            if value <= 0:
                continue
            index = value - 1
            if index < len(headers):
                return index, headers[index] if headers else ""

        if headers:
            return len(headers) - 1, headers[-1]
        return 0, ""

    def write_column(
        self,
        file_path: str,
        sheet_mode: str,
        sheet_name: str,
        text_content: str,
        target_column: str,
        write_mode: str,
        enable_new_column: bool,
        new_column_title: str,
    ):
        path = Path(file_path).expanduser()
        if path.suffix.lower() not in {".csv", ".xlsx", ".xlsm"}:
            path = path.with_suffix(".xlsx")

        _ensure_parent_dir(path)

        suffix = path.suffix.lower()
        headers: List[str] = []
        matrix: List[List[Any]] = []

        # 处理工作表模式
        if sheet_mode == "新建内页":
            # 新建内页模式：不读取现有数据，创建空表格
            if not path.exists():
                # 文件不存在，创建新文件
                pass
            # 对于新建内页，我们始终从空表格开始
            headers = []
            matrix = []
        else:
            # 使用现有模式：读取指定工作表的数据
            if path.exists():
                if suffix == ".csv":
                    headers, dict_rows = _read_csv_table(path, 0)
                else:
                    headers, dict_rows = _read_xlsx_table(path, sheet_name or None, 0)
                matrix = _table_to_matrix(headers, dict_rows)

        normalized_text = _normalize_newlines(text_content or "")
        lines = [] if normalized_text == "" else normalized_text.split("\n")
        line_count = len(lines)

        column_index, matched_title = self._resolve_column_index(headers, target_column)

        if enable_new_column:
            base_index = column_index if headers else 0
            insert_index = min(base_index + (1 if headers else 0), len(headers))
            column_title = new_column_title.strip() or (matched_title or f"列{insert_index + 1}")
            headers.insert(insert_index, column_title)
            for row in matrix:
                row.insert(insert_index, "")
            column_index = insert_index
        else:
            if not headers:
                headers.append(new_column_title.strip() or "列1")
                column_index = 0
            while len(headers) <= column_index:
                headers.append(f"列{len(headers) + 1}")
                for row in matrix:
                    row.append("")

        _ensure_table_dimensions(headers, matrix, len(matrix))

        legacy_write_map = {
            "覆盖": "覆盖重写",
            "填充": "空位追加",
        }

        mode_choice = legacy_write_map.get(write_mode, write_mode or "覆盖重写")

        if mode_choice == "覆盖重写":
            target_rows = max(len(matrix), line_count)
            _ensure_table_dimensions(headers, matrix, target_rows)
            for row_idx in range(len(matrix)):
                value = lines[row_idx] if row_idx < line_count else ""
                matrix[row_idx][column_index] = value
        elif mode_choice == "空位追加":
            start_row = 0
            for row_idx in range(len(matrix)):
                if not matrix[row_idx][column_index]:
                    start_row = row_idx
                    break
            else:
                start_row = len(matrix)
            required_rows = start_row + line_count
            _ensure_table_dimensions(headers, matrix, required_rows)
            for offset, value in enumerate(lines):
                row_idx = start_row + offset
                matrix[row_idx][column_index] = value
        else:
            raise ValueError("write_mode 仅支持 覆盖重写 或 空位追加")

        dict_rows = _matrix_to_dict_rows(headers, matrix)
        try:
            if suffix == ".csv":
                _write_csv_table(path, headers, dict_rows)
            else:
                with _WorkbookSheetManager(path, sheet_mode, sheet_name) as (wb, ws):
                    ws.delete_rows(1, ws.max_row)
                    ws.append(headers)
                    for row in dict_rows:
                        ws.append([row.get(header, "") for header in headers])
        except Exception as exc:  # pragma: no cover
            return (f"写入失败: {exc}", 0, text_content)

        return (str(path), line_count, text_content)
    
    def _load_workbook_with_sheet_mode(self, path: Path, sheet_mode: str, sheet_name: str):
        """
        根据工作表模式加载或创建工作表
        """
        if Workbook is None:
            raise RuntimeError("需要 openpyxl 才能写入 XLSX 表格")
        
        if path.exists():
            wb = load_workbook(path)
        else:
            wb = Workbook()
            # 删除默认工作表，因为我们要创建新的
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
        
        try:
            if sheet_mode == "新建内页":
                # 新建内页模式
                if sheet_name and sheet_name.strip():
                    # 使用指定名称
                    new_sheet_name = sheet_name.strip()
                    # 确保名称唯一
                    counter = 1
                    original_name = new_sheet_name
                    while new_sheet_name in wb.sheetnames:
                        new_sheet_name = f"{original_name}_{counter}"
                        counter += 1
                    ws = wb.create_sheet(title=new_sheet_name)
                else:
                    # 自动生成名称: Sheet1, Sheet2, ...
                    sheet_counter = len(wb.sheetnames) + 1
                    new_sheet_name = f"Sheet{sheet_counter}"
                    # 确保名称唯一
                    while new_sheet_name in wb.sheetnames:
                        sheet_counter += 1
                        new_sheet_name = f"Sheet{sheet_counter}"
                    ws = wb.create_sheet(title=new_sheet_name)
                print(f"✅ 创建新工作表: {ws.title}")
            else:
                # 使用现有模式
                if sheet_name and sheet_name.strip():
                    # 使用指定工作表
                    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
                else:
                    # 使用第一个工作表
                    ws = wb.active if wb.active else wb.create_sheet(title="Sheet1")
            
            yield wb, ws
        finally:
            wb.save(path)
            wb.close()


class buding_随机行输出:
    CATEGORY = "Buding/TextStory/文本处理"
    RETURN_TYPES = ("STRING", "BWF_TEXT_LIST", "INT", "INT")
    RETURN_NAMES = ("随机文本", "随机列表", "实际数量", "使用种子")
    FUNCTION = "sample"

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                # 接受字符串类型，支持上游连接或直接输入多行文本
                "lines": ("STRING", {"multiline": True, "default": "", "tooltip": "可输入多行文本，或连接上游 STRING/TEXT 节点"}),
                # 抽取数量上限提高到 300
                "sample_count": ("INT", {"default": 1, "min": 1, "max": 300, "tooltip": "抽取条目数量（上限300）"}),
                "unique": ("BOOLEAN", {"default": True, "tooltip": "开启后避免重复"}),
                "randomize": ("BOOLEAN", {"default": True, "tooltip": "开启后自动生成随机种子"}),
                "seed": ("INT", {"default": 0, "tooltip": "随机种子，randomize 关闭时生效"}),
                # fallback_to_all：不足时返回全部可用条目（例如上游1000行会返回全部1000行），关闭则报错
                "fallback_to_all": ("BOOLEAN", {"default": True, "tooltip": "数量不足时返回全部可用条目；关闭则抛出数量不足错误"}),
            }
        }

    def sample(
        self,
        lines: Any,
        sample_count: int,
        unique: bool,
        randomize: bool,
        seed: int,
        fallback_to_all: bool,
    ):
        items = [line for line in _string_to_list(lines) if line]
        if not items:
            return ("", [], 0, 0)

        if randomize or seed == 0:
            used_seed = random.randint(1, 2_147_483_647)
        else:
            used_seed = seed

        rng = random.Random(used_seed)

        if unique:
            population = list(dict.fromkeys(items))
            if sample_count > len(population):
                if fallback_to_all:
                    selected = population
                else:
                    raise ValueError("可用条目数量不足以满足抽取数量")
            else:
                selected = rng.sample(population, sample_count)
        else:
            population = items
            selected = [rng.choice(population) for _ in range(sample_count)]

        return ("\n".join(selected), selected, len(selected), used_seed)


class buding_文本合并器:
    CATEGORY = "Buding/TextStory/文本处理"
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("合并文本",)
    FUNCTION = "merge_texts"

    @classmethod
    def INPUT_TYPES(cls):
        inputs = {"required": {}}
        for i in range(1, 11):
            inputs["required"][f"text{i}"] = ("STRING", {"default": "", "tooltip": f"输入第{i}段文本，可留空"})
        inputs["required"]["delimiter"] = ("STRING", {"default": "", "multiline": True, "tooltip": "合并时使用的分隔符，留空默认换行"})
        inputs["required"]["enable_random"] = ("BOOLEAN", {"default": False, "tooltip": "开启后随机打乱文本顺序"})
        inputs["required"]["seed"] = ("INT", {"default": 0, "min": 0, "max": 0xFFFFFFFFFFFFFFFF, "tooltip": "随机种子，0 表示自动"})
        return inputs

    def merge_texts(
        self,
        text1,
        text2,
        text3,
        text4,
        text5,
        text6,
        text7,
        text8,
        text9,
        text10,
        delimiter,
        enable_random,
        seed,
    ):
        texts = [text1, text2, text3, text4, text5, text6, text7, text8, text9, text10]
        non_empty = [t for t in texts if str(t).strip()]

        delimiter = delimiter if str(delimiter).strip() else "\n"

        if enable_random:
            rng = random.Random()
            if seed == 0:
                rng.seed()
            else:
                rng.seed(seed)
            rng.shuffle(non_empty)

        merged = delimiter.join(non_empty)
        return (merged,)


class buding_标题拆分器:
    CATEGORY = "Buding/TextStory/文本处理"
    RETURN_TYPES = tuple(["STRING"] * 15)
    RETURN_NAMES = tuple(f"标题{i}" for i in range(1, 16))
    FUNCTION = "split_titles"

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "text": ("STRING", {"default": "", "multiline": True, "tooltip": "包含多个标题与正文的文本"}),
                "title_prefix": ("STRING", {"default": "一.", "tooltip": "用于识别标题的前缀，可用 | 分隔多个规则"}),
            }
        }

    def split_titles(self, text, title_prefix):
        lines = _normalize_newlines(text or "").split("\n")
        segments = []
        current_title = None
        current_content: List[str] = []

        prefixes = [p.strip() for p in str(title_prefix).split('|') if p.strip()]
        if not prefixes:
            prefixes = ["一."]

        for line in lines:
            stripped = line.strip()
            matched = False
            for prefix in prefixes:
                match = re.match(r"^(" + re.escape(prefix) + r".*)$", stripped)
                if match:
                    if current_title is not None:
                        segments.append((current_title, "\n".join(current_content).strip()))
                    current_title = match.group(1)
                    current_content = []
                    matched = True
                    break
            if not matched:
                current_content.append(line)

        if current_title is not None:
            segments.append((current_title, "\n".join(current_content).strip()))

        outputs: List[str] = []
        for title, content in segments[:15]:
            outputs.append(f"{{{title}}}\n{content}")

        while len(outputs) < 15:
            outputs.append("")

        return tuple(outputs)


class buding_剧情故事生成:
     STORY_TYPES = ["成长励志故事", "婚恋家庭故事", "现实共鸣故事", "悬疑反转故事", "燃爽逆袭故事", "古风玄幻故事", "科幻赛博故事", "都市职场故事", "古偶虐恋故事", "甜宠治愈故事", "金羊毛型故事", "伙伴之情型故事", "如愿以偿型故事", "麻烦家伙型故事", "变迁仪式型故事", "愚者成功型故事", "推理侦探型故事", "鬼屋型故事", "超级英雄型故事", "被制度化型故事"]
     STRUCTURE_TYPES = ["经典三幕式结构", "起承转合结构", "英雄之旅结构", "环形结构", "线性叙事结构", "多线叙事结构", "麦基人物线结构", "费希特曲线结构"]

     @classmethod
     def INPUT_TYPES(s):
         return {
             "required": {
                 "theme": ("STRING", {"default": "修真", "tooltip": "故事主题，可用顿号分隔多个候选值"}),
                 "topic": ("STRING", {"default": "都市人孤独感治愈", "tooltip": "故事选题或方向，可用顿号分隔多个候选"}),
                 "protagonist_name": ("STRING", {"default": "DD布丁", "tooltip": "主角姓名，支持多个候选"}),
                 "protagonist_background": ("STRING", {"default": "散修", "tooltip": "主角背景设定或身份描述"}),
                 "motivation": ("STRING", {"default": "寻找意义", "tooltip": "主角行动动机或核心驱动力"}),
                 "desire": ("STRING", {"default": "被爱与理解", "tooltip": "主角核心欲望或目标"}),
                 "conflict": ("STRING", {"default": "工作压力与人际疏离", "tooltip": "故事主要冲突点"}),
                 "action": ("STRING", {"default": "寻求帮助", "tooltip": "主角采取的关键行动"}),
                 "obstacles": ("STRING", {"default": "自我怀疑", "tooltip": "主角遇到的障碍或阻力"}),
                 "result": ("STRING", {"default": "获得友谊", "tooltip": "故事结局或阶段性成果"}),
                 "positive_value": ("STRING", {"default": "善良", "tooltip": "想要强调的正向价值观"}),
                 "negative_value": ("STRING", {"default": "自私", "tooltip": "需要规避或反思的负向价值"}),
                 "story_type": (s.STORY_TYPES, {"default": "成长励志", "tooltip": "故事类型，可选预设或自定义"}),
                 "structure_type": (s.STRUCTURE_TYPES, {"default": "经典三幕式", "tooltip": "叙事结构，可选预设或自定义"}),
                 "batch_count": ("INT", {"default": 1, "min": 1, "max": 10, "tooltip": "一次生成的故事数量"}),
                 "enable_random_story_type": ("BOOLEAN", {"default": False, "tooltip": "开启后在预设故事类型间随机选择"}),
                 "enable_random_structure_type": ("BOOLEAN", {"default": False, "tooltip": "开启后在预设叙事结构间随机选择"}),
                 "enable_trending": ("BOOLEAN", {"default": False, "tooltip": "开启后提示融入爆款元素"}),
                 "trending_element1": ("STRING", {"default": "", "tooltip": "爆款元素候选 1，支持顿号分隔"}),
                 "trending_element2": ("STRING", {"default": "", "tooltip": "爆款元素候选 2"}),
                 "trending_element3": ("STRING", {"default": "", "tooltip": "爆款元素候选 3"}),
                 "enable_custom_story_type": ("BOOLEAN", {"default": False, "tooltip": "开启后优先使用自定义故事类型"}),
                 "custom_story_type": ("STRING", {"default": "", "tooltip": "自定义故事类型内容"}),
                 "enable_custom_structure_type": ("BOOLEAN", {"default": False, "tooltip": "开启后优先使用自定义叙事结构"}),
                 "custom_structure_type": ("STRING", {"default": "", "tooltip": "自定义叙事结构内容"}),
                 "dialogue_percentage": ("INT", {"default": 60, "min": 0, "max": 100, "tooltip": "预期的台词占比（百分比）"}),
                 "word_count": ("STRING", {"default": "300-500字", "tooltip": "目标字数范围或固定字数"}),
                 "seed": ("INT", {"default": 0, "min": 0, "max": 999999, "tooltip": "随机种子，0 表示自动"})
             }
         }

     RETURN_TYPES = tuple(["STRING"] * 10)  # Max 10 for batch
     RETURN_NAMES = tuple(f"story_{i+1}" for i in range(10))
     FUNCTION = "generate"
     CATEGORY = "Buding"

     def generate(self, theme, topic, protagonist_name, protagonist_background, motivation, desire, conflict, action, obstacles, result, positive_value, negative_value, story_type, structure_type, batch_count, enable_random_story_type, enable_random_structure_type, enable_trending, trending_element1, trending_element2, trending_element3, enable_custom_story_type, custom_story_type, enable_custom_structure_type, custom_structure_type, dialogue_percentage, word_count, seed):
         def random_select(text):
             if '、' in text:
                 options = [t.strip() for t in text.split('、') if t.strip()]
                 if options:
                     return random.choice(options)
             return text
         
         prompts = []
         for i in range(batch_count):
             if seed == 0:
                 random.seed(int(time.time() * 1000000) % 1000000 + i)
             else:
                 random.seed(seed + i)
             
             theme_s = random_select(theme)
             topic_s = random_select(topic)
             protagonist_name_s = random_select(protagonist_name)
             protagonist_background_s = random_select(protagonist_background)
             motivation_s = random_select(motivation)
             desire_s = random_select(desire)
             conflict_s = random_select(conflict)
             action_s = random_select(action)
             obstacles_s = random_select(obstacles)
             result_s = random_select(result)
             positive_value_s = random_select(positive_value)
             negative_value_s = random_select(negative_value)
             word_count_s = random_select(word_count)
             
             selected_story_type = custom_story_type if enable_custom_story_type and custom_story_type else (random.choice(self.STORY_TYPES) if enable_random_story_type else story_type)
             selected_structure_type = custom_structure_type if enable_custom_structure_type and custom_structure_type else (random.choice(self.STRUCTURE_TYPES) if enable_random_structure_type else structure_type)
             
             prompt_parts = [f"创作一个{selected_story_type}，使用{selected_structure_type}。"]
             if theme_s:
                 prompt_parts.append(f"主题：{theme_s}，")
             if topic_s:
                 prompt_parts.append(f"选题：{topic_s}。")
             if protagonist_name_s or protagonist_background_s:
                 bg = f"（背景：{protagonist_background_s}）" if protagonist_background_s else ""
                 prompt_parts.append(f"主角：{protagonist_name_s}{bg}，")
             if motivation_s:
                 prompt_parts.append(f"动机：{motivation_s}，")
             if desire_s:
                 prompt_parts.append(f"欲望：{desire_s}，")
             if conflict_s:
                 prompt_parts.append(f"冲突：{conflict_s}，")
             if action_s:
                 prompt_parts.append(f"动作：{action_s}，")
             if obstacles_s:
                 prompt_parts.append(f"障碍：{obstacles_s}，")
             if result_s:
                 prompt_parts.append(f"结果：{result_s}，")
             if positive_value_s:
                 prompt_parts.append(f"正价值观：{positive_value_s}，")
             if negative_value_s:
                 prompt_parts.append(f"负价值观：{negative_value_s}。")
             if dialogue_percentage != 60 or word_count_s != "300-500字":
                 prompt_parts.append(f"角色台词占比：{dialogue_percentage}% ，故事字数：{word_count_s} 。")
             prompt_parts.append("融入编剧要素，确保完整性与张力。")
             
             prompt = "".join(prompt_parts)
             if enable_trending:
                 elements = [e for e in [trending_element1, trending_element2, trending_element3] if e]
                 if elements:
                     prompt += f" 【爆款元素融入，请选择1-3个爆款元素融入到故事中】{{{','.join(elements)}}}"
             prompts.append(prompt)
         
         # Pad to 10 outputs
         while len(prompts) < 10:
             prompts.append("")
         return tuple(prompts)


class buding_爆款元素组合:
    CATEGORY = "Buding/TextStory/热点组合"
    RETURN_TYPES = tuple(["STRING"] * 10)
    RETURN_NAMES = tuple(f"爆款元素{i}" for i in range(1, 11))
    FUNCTION = "combine"

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "tags": ("STRING", {"default": "卡点、变装、BGM", "tooltip": "标签库，顿号分隔"}),
                "passwords": ("STRING", {"default": "JK、乡土青年、逗比", "tooltip": "密码库，顿号分隔"}),
                "count": ("INT", {"default": 3, "min": 1, "max": 10, "tooltip": "需要组合的数量"}),
                "seed": ("INT", {"default": 0, "min": 0, "max": 999999, "tooltip": "随机种子，0 表示自动"}),
            }
        }

    def combine(self, tags, passwords, count, seed):
        random.seed(seed if seed != 0 else int(time.time() * 1000000) % 1_000_000)
        tag_list = [t.strip() for t in str(tags).split('、') if t.strip()]
        pw_list = [p.strip() for p in str(passwords).split('、') if p.strip()]
        if not tag_list and not pw_list:
            results = [""] * count
        else:
            results = []
            for _ in range(count):
                if tag_list and not pw_list:
                    combined = random.choice(tag_list)
                elif not tag_list and pw_list:
                    combined = random.choice(pw_list)
                else:
                    combined = random.choice(tag_list) + random.choice(pw_list)
                results.append(combined)
        while len(results) < 10:
            results.append("")
        return tuple(results)


class buding_爆款组合简版:
    CATEGORY = "Buding/TextStory/热点组合"
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("组合元素",)
    FUNCTION = "combine"

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "tags": ("STRING", {"default": "卡点、变装、BGM", "tooltip": "标签库，顿号分隔"}),
                "passwords": ("STRING", {"default": "JK、乡土青年、逗比", "tooltip": "密码库，顿号分隔"}),
                "seed": ("INT", {"default": 0, "min": 0, "max": 999999, "tooltip": "随机种子，0 表示自动"}),
            }
        }

    def combine(self, tags, passwords, seed):
        random.seed(seed if seed != 0 else int(time.time() * 1000000) % 1_000_000)
        tag_list = [t.strip() for t in str(tags).split('、') if t.strip()]
        pw_list = [p.strip() for p in str(passwords).split('、') if p.strip()]
        if not tag_list and not pw_list:
            return ("",)
        if tag_list and not pw_list:
            return (random.choice(tag_list),)
        if not tag_list and pw_list:
            return (random.choice(pw_list),)
        return (random.choice(tag_list) + random.choice(pw_list),)


class buding_随机词条挑选:
    CATEGORY = "Buding/TextStory/热点组合"
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("随机结果",)
    FUNCTION = "pick"

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "items": ("STRING", {"default": "选项1、选项2、选项3", "tooltip": "候选列表，顿号分隔"}),
                "seed": ("INT", {"default": 0, "min": 0, "max": 999999, "tooltip": "随机种子，0 表示自动"}),
            }
        }

    def pick(self, items, seed):
        if seed == 0:
            random.seed(int.from_bytes(os.urandom(4), 'big'))
        else:
            random.seed(seed)
        item_list = [i.strip() for i in str(items).split('、') if i.strip()]
        if not item_list:
            return ("",)
        selected = random.choice(item_list)
        return (selected,)


class buding_短视频脚本生成:
    CATEGORY = "Buding/TextStory/剧情创作"
    RETURN_TYPES = tuple(["STRING"] * 10)
    RETURN_NAMES = tuple(f"短视频脚本{i}" for i in range(1, 11))
    FUNCTION = "generate"

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "topic": ("STRING", {"default": "都市人孤独感治愈", "tooltip": "短视频主题，可用顿号分隔多个候选"}),
                "style": ("STRING", {"default": "", "tooltip": "风格设定，可选"}),
                "hook": ("STRING", {"default": "", "tooltip": "开头吸引点，可选"}),
                "conflict": ("STRING", {"default": "", "tooltip": "故事冲突，可选"}),
                "twist": ("STRING", {"default": "", "tooltip": "剧情反转，可选"}),
                "golden_sentence": ("STRING", {"default": "", "tooltip": "金句或亮点，可选"}),
                "duration": ("INT", {"default": 30, "min": 1, "max": 120, "tooltip": "时长（秒）"}),
                "audience": ("STRING", {"default": "年轻人", "tooltip": "目标受众"}),
                "batch_count": ("INT", {"default": 1, "min": 1, "max": 10, "tooltip": "生成脚本数量"}),
                "enable_trending": ("BOOLEAN", {"default": False, "tooltip": "是否提示融入爆款元素"}),
                "trending_element1": ("STRING", {"default": "", "tooltip": "爆款元素候选 1"}),
                "trending_element2": ("STRING", {"default": "", "tooltip": "爆款元素候选 2"}),
                "trending_element3": ("STRING", {"default": "", "tooltip": "爆款元素候选 3"}),
                "interaction": ("STRING", {"default": "", "tooltip": "互动方式，可选"}),
                "cta": ("STRING", {"default": "", "tooltip": "行动号召，可选"}),
                "title_generation": ("STRING", {"default": "", "tooltip": "标题生成要求"}),
                "seed": ("INT", {"default": 0, "min": 0, "max": 999999, "tooltip": "随机种子，0 表示自动"}),
            }
        }

    def generate(
        self,
        topic,
        style,
        hook,
        conflict,
        twist,
        golden_sentence,
        duration,
        audience,
        batch_count,
        enable_trending,
        trending_element1,
        trending_element2,
        trending_element3,
        interaction,
        cta,
        title_generation,
        seed,
    ):
        def random_select(text: str) -> str:
            if '、' in text:
                options = [t.strip() for t in text.split('、') if t.strip()]
                if options:
                    return random.choice(options)
            return text

        prompts: List[str] = []
        for i in range(batch_count):
            if seed == 0:
                random.seed(int.from_bytes(os.urandom(4), 'big') + i)
            else:
                random.seed(seed + i)

            topic_s = random_select(topic)
            style_s = random_select(style)
            hook_s = random_select(hook)
            conflict_s = random_select(conflict)
            twist_s = random_select(twist)
            golden_sentence_s = random_select(golden_sentence)
            audience_s = random_select(audience)
            interaction_s = random_select(interaction)
            cta_s = random_select(cta)
            title_generation_s = random_select(title_generation)

            prompt_parts = [f"创作一个{topic_s}短视频片段："]
            if style_s:
                prompt_parts.append(f"风格{style_s}，")
            if hook_s:
                prompt_parts.append(f"开头钩子{hook_s}，")
            if conflict_s:
                prompt_parts.append(f"冲突{conflict_s}，")
            if twist_s:
                prompt_parts.append(f"转折{twist_s}。")
            if golden_sentence_s:
                prompt_parts.append(f"金句{golden_sentence_s}，")
            prompt_parts.append(f"时长{duration}秒，受众{audience_s}。")
            if interaction_s:
                prompt_parts.append(f"互动{interaction_s}，")
            if cta_s:
                prompt_parts.append(f"CTA{cta_s}。")

            trendings_list = [e for e in [trending_element1, trending_element2, trending_element3] if e]
            if enable_trending and trendings_list:
                num_select = random.randint(1, min(3, len(trendings_list)))
                selected = random.sample(trendings_list, num_select)
                trendings = ",".join(selected)
                prompt_parts.append(f"【爆款元素融入，请选择1-3个爆款元素融入到短视频中】{{{trendings}}}")

            if title_generation_s:
                prompt_parts.append(f"最后，生成5个适配{title_generation_s}的标题。")

            prompts.append("".join(prompt_parts))

        while len(prompts) < 10:
            prompts.append("")
        return tuple(prompts)


NODE_CLASS_MAPPINGS = {
    "buding_简易文本替换": buding_简易文本替换,
    "buding_符号文本转换": buding_符号文本转换,
    "buding_列文本转换": buding_列文本转换,
    "buding_自动文本转换": buding_自动文本转换,
    "buding_文本保存TXT": buding_文本保存TXT,
    "buding_读取TXT文件": buding_读取TXT文件,
    "buding_随机行输出": buding_随机行输出,
    "buding_读取表格数据": buding_读取表格数据,
    "buding_写入表格数据": buding_写入表格数据,
    "buding_文本合并器": buding_文本合并器,
    "buding_标题拆分器": buding_标题拆分器,
    "buding_剧情故事生成": buding_剧情故事生成,
    "buding_爆款元素组合": buding_爆款元素组合,
    "buding_爆款组合简版": buding_爆款组合简版,
    "buding_随机词条挑选": buding_随机词条挑选,
    "buding_短视频脚本生成": buding_短视频脚本生成,
}

NODE_DISPLAY_NAME_MAPPINGS = {name: name for name in NODE_CLASS_MAPPINGS}
